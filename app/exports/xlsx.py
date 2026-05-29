"""XLSX budget export generation and storage."""

from __future__ import annotations

from dataclasses import dataclass
from datetime import UTC, date, datetime
from decimal import Decimal
from hashlib import sha256
from io import BytesIO

import psycopg
from openpyxl import Workbook
from psycopg import Connection

from app.budget import load_category_budgets_in_connection
from app.sanitize import sanitize_error_message

XLSX_CONTENT_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
REQUIRED_SHEETS = (
    "Summary",
    "Category averages",
    "Monthly history",
    "Recurring expenses",
    "Excluded one-offs",
    "Raw transactions",
    "Rules",
    "Forecast assumptions",
)


@dataclass(frozen=True)
class ExportFile:
    id: int
    filename: str
    content_type: str
    content: bytes


@dataclass(frozen=True)
class ExportRunSummary:
    id: int
    export_type: str
    status: str
    filename: str | None
    file_id: int | None
    started_at: datetime | None
    finished_at: datetime | None
    error_message: str | None


def create_pending_budget_export(database_url: str, *, created_by: str | None = None) -> int:
    with psycopg.connect(database_url) as connection:
        with connection.transaction():
            year_month = _latest_year_month(connection)
            period_start = date.fromisoformat(f"{year_month}-01")
            period_end = _period_end(period_start)
            return _create_export_run(connection, period_start, period_end, created_by, status="pending")


def generate_budget_export(
    database_url: str,
    *,
    created_by: str | None = None,
    run_id: int | None = None,
    retention_days: int | None = None,
) -> int:
    _validate_retention_days(retention_days)
    with psycopg.connect(database_url, autocommit=True) as connection:
        prepared_run_id, year_month = _prepare_export_run(connection, created_by=created_by, run_id=run_id)
        try:
            content = _build_workbook(connection, year_month)
            filename = f"budget-averages-{year_month}.xlsx"
            with connection.transaction():
                _store_export_file(connection, prepared_run_id, filename, content)
                _finish_export_run(connection, prepared_run_id, "completed", None)
                _prune_old_export_runs(connection, retention_days)
        except Exception as exc:
            with connection.transaction():
                _finish_export_run(connection, prepared_run_id, "failed", sanitize_error_message(exc))
            raise
    return prepared_run_id


def list_export_runs(database_url: str) -> list[ExportRunSummary]:
    with psycopg.connect(database_url) as connection:
        rows = connection.execute(
            """
            SELECT
                export_runs.id,
                export_runs.export_type,
                export_runs.status,
                export_files.filename,
                export_files.id,
                export_runs.started_at,
                export_runs.finished_at,
                export_runs.error_message
            FROM export_runs
            LEFT JOIN export_files ON export_files.export_run_id = export_runs.id
            ORDER BY export_runs.id DESC
            LIMIT 20
            """
        ).fetchall()
    return [
        ExportRunSummary(
            id=_read_int(row[0]),
            export_type=str(row[1]),
            status=str(row[2]),
            filename=_optional_str(row[3]),
            file_id=_optional_int(row[4]),
            started_at=_optional_datetime(row[5]),
            finished_at=_optional_datetime(row[6]),
            error_message=_optional_str(row[7]),
        )
        for row in rows
    ]


def load_export_file(database_url: str, file_id: int) -> ExportFile | None:
    with psycopg.connect(database_url) as connection:
        row = connection.execute(
            """
            SELECT id, filename, content_type, content
            FROM export_files
            WHERE id = %s
            """,
            (file_id,),
        ).fetchone()
    if row is None:
        return None
    content = row[3]
    if not isinstance(content, bytes):
        raise RuntimeError(f"expected bytes content, got {type(content).__name__}")
    return ExportFile(
        id=_read_int(row[0]),
        filename=str(row[1]),
        content_type=str(row[2]),
        content=content,
    )


def _build_workbook(connection: Connection[tuple[object, ...]], year_month: str) -> bytes:
    workbook = Workbook()
    summary = workbook.active
    summary.title = REQUIRED_SHEETS[0]
    _write_summary_sheet(connection, summary, year_month)
    _write_category_averages_sheet(connection, workbook.create_sheet(REQUIRED_SHEETS[1]))
    _write_monthly_history_sheet(connection, workbook.create_sheet(REQUIRED_SHEETS[2]))
    _write_recurring_sheet(connection, workbook.create_sheet(REQUIRED_SHEETS[3]))
    _write_excluded_one_offs_sheet(connection, workbook.create_sheet(REQUIRED_SHEETS[4]))
    _write_raw_transactions_sheet(connection, workbook.create_sheet(REQUIRED_SHEETS[5]))
    _write_rules_sheet(connection, workbook.create_sheet(REQUIRED_SHEETS[6]))
    _write_assumptions_sheet(connection, workbook.create_sheet(REQUIRED_SHEETS[7]))

    stream = BytesIO()
    workbook.save(stream)
    return stream.getvalue()


def _write_summary_sheet(connection: Connection[tuple[object, ...]], sheet, year_month: str) -> None:
    row = connection.execute(
        """
        SELECT
            target_savings,
            safety_buffer,
            safe_to_spend,
            projected_savings,
            confidence
        FROM monthly_forecasts
        WHERE year_month = %s
        """,
        (year_month,),
    ).fetchone()
    sheet.append(["generated_at", datetime.now(UTC).isoformat()])
    sheet.append(["period covered", year_month])
    if row is None:
        sheet.append(["target savings", Decimal("0")])
        sheet.append(["safety buffer", Decimal("0")])
        sheet.append(["safe to spend", Decimal("0")])
        sheet.append(["projected savings", Decimal("0")])
        sheet.append(["forecast confidence", "low"])
        return
    sheet.append(["target savings", row[0]])
    sheet.append(["safety buffer", row[1]])
    sheet.append(["safe to spend", row[2]])
    sheet.append(["projected savings", row[3]])
    sheet.append(["forecast confidence", row[4]])


def _write_category_averages_sheet(connection: Connection[tuple[object, ...]], sheet) -> None:
    sheet.append(
        [
            "Category",
            "3M average",
            "6M average",
            "12M average",
            "Current month",
            "Suggested budget",
            "Included in forecast",
            "Notes",
        ]
    )
    for row in load_category_budgets_in_connection(connection):
        sheet.append(
            [
                row.category,
                row.average_3m,
                row.average_6m,
                row.average_12m,
                row.current_month,
                row.suggested_budget,
                row.included_in_forecast,
                f"Excluded this month: {row.excluded_from_forecast}",
            ]
        )


def _write_monthly_history_sheet(connection: Connection[tuple[object, ...]], sheet) -> None:
    sheet.append(["Month", "Income", "Fixed costs", "Variable costs", "Savings", "Excluded spending", "Net cashflow"])
    rows = connection.execute(
        """
        SELECT
            monthly_forecasts.year_month,
            monthly_forecasts.income_received,
            monthly_forecasts.fixed_costs_paid,
            monthly_forecasts.variable_spent,
            monthly_forecasts.projected_savings,
            COALESCE(sum(abs(raw_transactions.amount)) FILTER (
                WHERE enriched_transactions.is_excluded_from_budget = true
            ), 0),
            monthly_forecasts.projected_savings
        FROM monthly_forecasts
        LEFT JOIN raw_transactions ON to_char(raw_transactions.booking_date, 'YYYY-MM') = monthly_forecasts.year_month
        LEFT JOIN enriched_transactions ON enriched_transactions.raw_transaction_id = raw_transactions.id
        GROUP BY monthly_forecasts.id
        ORDER BY monthly_forecasts.year_month
        """
    ).fetchall()
    for row in rows:
        sheet.append(list(row))


def _write_recurring_sheet(connection: Connection[tuple[object, ...]], sheet) -> None:
    sheet.append(["Name", "Category", "Cadence", "Expected amount", "Next expected date", "Confidence", "Confirmed"])
    rows = connection.execute(
        """
        SELECT
            recurring_series.name,
            COALESCE(categories.name, 'Unknown'),
            recurring_series.cadence,
            recurring_series.expected_amount,
            recurring_series.next_expected_date,
            recurring_series.confidence,
            recurring_series.is_confirmed
        FROM recurring_series
        LEFT JOIN categories ON categories.id = recurring_series.category_id
        ORDER BY recurring_series.name
        """
    ).fetchall()
    for row in rows:
        sheet.append(list(row))


def _write_excluded_one_offs_sheet(connection: Connection[tuple[object, ...]], sheet) -> None:
    sheet.append(["Date", "Amount", "Merchant", "Description", "Category", "Reason"])
    rows = connection.execute(
        """
        SELECT
            raw_transactions.booking_date,
            raw_transactions.amount,
            COALESCE(merchants.name, raw_transactions.counterparty_name, ''),
            raw_transactions.description,
            COALESCE(categories.name, 'Unknown'),
            enriched_transactions.notes
        FROM enriched_transactions
        JOIN raw_transactions ON raw_transactions.id = enriched_transactions.raw_transaction_id
        LEFT JOIN merchants ON merchants.id = enriched_transactions.merchant_id
        LEFT JOIN categories ON categories.id = enriched_transactions.category_id
        WHERE enriched_transactions.is_excluded_from_budget = true
            OR enriched_transactions.is_one_off = true
        ORDER BY raw_transactions.booking_date
        """
    ).fetchall()
    for row in rows:
        sheet.append(list(row))


def _write_raw_transactions_sheet(connection: Connection[tuple[object, ...]], sheet) -> None:
    sheet.append(["Date", "Amount", "Counterparty", "Description", "Category", "Classification", "Needs review"])
    rows = connection.execute(
        """
        SELECT
            raw_transactions.booking_date,
            raw_transactions.amount,
            raw_transactions.counterparty_name,
            raw_transactions.description,
            COALESCE(categories.name, 'Unknown'),
            COALESCE(enriched_transactions.classification_method, 'none'),
            enriched_transactions.needs_review
        FROM raw_transactions
        LEFT JOIN enriched_transactions ON enriched_transactions.raw_transaction_id = raw_transactions.id
        LEFT JOIN categories ON categories.id = enriched_transactions.category_id
        ORDER BY raw_transactions.booking_date
        """
    ).fetchall()
    for row in rows:
        sheet.append(list(row))


def _write_rules_sheet(connection: Connection[tuple[object, ...]], sheet) -> None:
    sheet.append(["Name", "Priority", "Active", "Field", "Operator", "Pattern", "Category", "Merchant"])
    rows = connection.execute(
        """
        SELECT
            categorization_rules.name,
            categorization_rules.priority,
            categorization_rules.is_active,
            categorization_rules.field,
            categorization_rules.operator,
            categorization_rules.pattern,
            categories.name,
            merchants.name
        FROM categorization_rules
        LEFT JOIN categories ON categories.id = categorization_rules.category_id
        LEFT JOIN merchants ON merchants.id = categorization_rules.merchant_id
        ORDER BY categorization_rules.priority, categorization_rules.id
        """
    ).fetchall()
    for row in rows:
        sheet.append(list(row))


def _write_assumptions_sheet(connection: Connection[tuple[object, ...]], sheet) -> None:
    sheet.append(["Key", "Value"])
    rows = connection.execute("SELECT key, value_json FROM app_settings ORDER BY key").fetchall()
    if not rows:
        sheet.append(["source", "deterministic sample or current database state"])
        return
    for row in rows:
        sheet.append([row[0], str(row[1])])


def _latest_year_month(connection: Connection[tuple[object, ...]]) -> str:
    row = connection.execute("SELECT year_month FROM monthly_forecasts ORDER BY year_month DESC LIMIT 1").fetchone()
    if row is None:
        return date.today().strftime("%Y-%m")
    return str(row[0])


def _period_end(period_start: date) -> date:
    next_month = (
        period_start.replace(year=period_start.year + 1, month=1)
        if period_start.month == 12
        else period_start.replace(month=period_start.month + 1)
    )
    return date.fromordinal(next_month.toordinal() - 1)


def _create_export_run(
    connection: Connection[tuple[object, ...]],
    period_start: date,
    period_end: date,
    created_by: str | None,
    *,
    status: str,
) -> int:
    row = connection.execute(
        """
        INSERT INTO export_runs (
            export_type,
            period_start,
            period_end,
            status,
            started_at,
            created_by,
            metadata_json
        )
        VALUES (
            'budget_averages',
            %s,
            %s,
            %s,
            CASE WHEN %s = 'running' THEN now() ELSE NULL END,
            %s,
            '{}'::jsonb
        )
        RETURNING id
        """,
        (period_start, period_end, status, status, created_by),
    ).fetchone()
    if row is None:
        raise RuntimeError("export run insert did not return an id")
    return _read_int(row[0])


def _prepare_export_run(
    connection: Connection[tuple[object, ...]],
    *,
    created_by: str | None,
    run_id: int | None,
) -> tuple[int, str]:
    with connection.transaction():
        if run_id is None:
            year_month = _latest_year_month(connection)
            period_start = date.fromisoformat(f"{year_month}-01")
            period_end = _period_end(period_start)
            prepared_run_id = _create_export_run(
                connection,
                period_start,
                period_end,
                created_by,
                status="running",
            )
            return prepared_run_id, year_month

        period_start = _start_export_run(connection, run_id, created_by)
        return run_id, period_start.strftime("%Y-%m")


def _start_export_run(
    connection: Connection[tuple[object, ...]],
    run_id: int,
    created_by: str | None,
) -> date:
    connection.execute("DELETE FROM export_files WHERE export_run_id = %s", (run_id,))
    row = connection.execute(
        """
        UPDATE export_runs
        SET
            status = 'running',
            started_at = now(),
            finished_at = NULL,
            error_message = NULL,
            created_by = COALESCE(created_by, %s)
        WHERE id = %s
        RETURNING period_start
        """,
        (created_by, run_id),
    ).fetchone()
    if row is None:
        raise ValueError(f"export run does not exist: {run_id}")
    period_start = row[0]
    if not isinstance(period_start, date):
        raise RuntimeError(f"expected date period_start, got {type(period_start).__name__}")
    return period_start


def _store_export_file(connection: Connection[tuple[object, ...]], run_id: int, filename: str, content: bytes) -> None:
    connection.execute(
        """
        INSERT INTO export_files (
            export_run_id,
            filename,
            content_type,
            content,
            size_bytes,
            sha256
        )
        VALUES (%s, %s, %s, %s, %s, %s)
        """,
        (run_id, filename, XLSX_CONTENT_TYPE, content, len(content), sha256(content).hexdigest()),
    )


def _finish_export_run(
    connection: Connection[tuple[object, ...]],
    run_id: int,
    status: str,
    error_message: str | None,
) -> None:
    connection.execute(
        """
        UPDATE export_runs
        SET status = %s, finished_at = now(), error_message = %s
        WHERE id = %s
        """,
        (status, error_message, run_id),
    )


def _validate_retention_days(retention_days: int | None) -> None:
    if retention_days is not None and retention_days < 1:
        raise ValueError("retention_days must be at least 1")


def _prune_old_export_runs(connection: Connection[tuple[object, ...]], retention_days: int | None) -> None:
    if retention_days is None:
        return
    connection.execute(
        """
        DELETE FROM export_runs
        WHERE status IN ('completed', 'failed')
            AND finished_at IS NOT NULL
            AND finished_at < now() - (%s * interval '1 day')
        """,
        (retention_days,),
    )


def _read_int(value: object) -> int:
    if not isinstance(value, int):
        raise RuntimeError(f"expected integer, got {type(value).__name__}")
    return value


def _optional_int(value: object) -> int | None:
    if value is None:
        return None
    return _read_int(value)


def _optional_str(value: object) -> str | None:
    if value is None:
        return None
    return str(value)


def _optional_datetime(value: object) -> datetime | None:
    if value is None:
        return None
    if not isinstance(value, datetime):
        raise RuntimeError(f"expected datetime, got {type(value).__name__}")
    return value
