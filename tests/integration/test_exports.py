from __future__ import annotations

import re
from collections.abc import Iterator
from hashlib import sha256
from io import BytesIO

import psycopg
import pytest
from flask import Flask
from openpyxl import load_workbook
from testcontainers.postgres import PostgresContainer

from app.config import AppConfig
from app.exports.xlsx import REQUIRED_SHEETS, XLSX_CONTENT_TYPE, generate_budget_export, load_export_file
from app.jobs.worker import run_until_drained
from app.migrate import run_migrations
from app.sample_data import load_sample_data
from app.web import create_app


@pytest.fixture
def sample_app() -> Iterator[Flask]:
    with PostgresContainer(
        image="postgres:16-alpine",
        username="treasuri",
        password="treasuri",
        dbname="treasuri",
        driver=None,
    ) as postgres:
        database_url = postgres.get_connection_url(driver=None)
        run_migrations(database_url)
        load_sample_data(database_url)
        yield create_app(
            AppConfig(
                app_env="test",
                secret_key="test-secret",
                database_url=database_url,
                allowed_emails=("dev-user@example.test",),
                oidc_enabled=False,
                oidc_testing_profile={
                    "sub": "dev-user",
                    "email": "dev-user@example.test",
                },
                oidc_cookie_secure=False,
                llm_enabled=False,
            ),
            {"TESTING": True},
        )


def test_generate_budget_export_stores_required_xlsx_sheets(sample_app: Flask) -> None:
    run_id = generate_budget_export(sample_app.config["DATABASE_URL"], created_by="dev-user@example.test")

    with psycopg.connect(sample_app.config["DATABASE_URL"]) as connection:
        row = connection.execute(
            """
            SELECT
                export_runs.status,
                export_files.id,
                export_files.filename,
                export_files.size_bytes,
                export_files.sha256,
                export_files.content
            FROM export_runs
            JOIN export_files ON export_files.export_run_id = export_runs.id
            WHERE export_runs.id = %s
            """,
            (run_id,),
        ).fetchone()

    assert row is not None
    assert row[0] == "completed"
    assert row[2] == "budget-averages-2026-05.xlsx"
    assert row[3] > 0
    assert row[4] == sha256(row[5]).hexdigest()

    export_file = load_export_file(sample_app.config["DATABASE_URL"], row[1])
    assert export_file is not None
    assert sha256(export_file.content).hexdigest() == row[4]
    workbook = load_workbook(BytesIO(export_file.content), read_only=True)

    assert workbook.sheetnames == list(REQUIRED_SHEETS)
    assert workbook["Summary"]["A1"].value == "generated_at"
    assert workbook["Category averages"]["A1"].value == "Category"
    assert workbook["Category averages"]["A2"].value == "Car"
    assert workbook["Recurring expenses"]["A1"].value == "Name"
    assert workbook["Recurring expenses"]["A2"].value == "Sample Streaming"
    assert workbook["Recurring expenses"]["B2"].value == "Subscriptions"
    assert workbook["Recurring expenses"]["C2"].value == "monthly"
    assert workbook["Raw transactions"]["A1"].value == "Date"


def test_export_routes_generate_and_download_postgres_blob(sample_app: Flask) -> None:
    client = sample_app.test_client()
    export_page = client.get("/export")
    csrf_token = _extract_csrf(export_page.get_data(as_text=True))

    generate_response = client.post(
        "/export/generate",
        data={"csrf_token": csrf_token},
        follow_redirects=True,
    )

    assert generate_response.status_code == 200
    assert b"pending" in generate_response.data

    with psycopg.connect(sample_app.config["DATABASE_URL"]) as connection:
        pending_row = connection.execute(
            """
            SELECT
                export_runs.id,
                export_runs.status,
                export_runs.created_by,
                pgqueuer.entrypoint,
                (convert_from(pgqueuer.payload, 'UTF8')::jsonb ->> 'run_id')::bigint
            FROM export_runs
            JOIN pgqueuer ON (convert_from(pgqueuer.payload, 'UTF8')::jsonb ->> 'run_id')::bigint = export_runs.id
            ORDER BY export_runs.id DESC
            LIMIT 1
            """
        ).fetchone()

    assert pending_row is not None
    assert isinstance(pending_row[0], int)
    assert pending_row[1:] == ("pending", "dev-user@example.test", "generate_xlsx_export", pending_row[0])

    run_until_drained(sample_app.config["APP_CONFIG"])

    completed_response = client.get("/export")
    assert completed_response.status_code == 200
    assert b"budget-averages-2026-05.xlsx" in completed_response.data

    match = re.search(rb'href="/export/files/(\d+)"', completed_response.data)
    assert match is not None

    download_response = client.get(f"/export/files/{match.group(1).decode()}")

    with psycopg.connect(sample_app.config["DATABASE_URL"]) as connection:
        row = connection.execute(
            """
            SELECT export_runs.created_by, export_files.sha256, export_files.content
            FROM export_files
            JOIN export_runs ON export_runs.id = export_files.export_run_id
            WHERE export_files.id = %s
            """,
            (int(match.group(1).decode()),),
        ).fetchone()

    assert download_response.status_code == 200
    assert download_response.headers["Content-Type"] == XLSX_CONTENT_TYPE
    assert "budget-averages-2026-05.xlsx" in download_response.headers["Content-Disposition"]
    assert row is not None
    assert row[0] == "dev-user@example.test"
    assert download_response.data == row[2]
    assert sha256(download_response.data).hexdigest() == row[1]
    workbook = load_workbook(BytesIO(download_response.data), read_only=True)
    assert workbook.sheetnames == list(REQUIRED_SHEETS)


def test_export_download_returns_404_for_unknown_file(sample_app: Flask) -> None:
    response = sample_app.test_client().get("/export/files/99999")

    assert response.status_code == 404


def test_export_download_requires_allowed_user(sample_app: Flask) -> None:
    run_id = generate_budget_export(sample_app.config["DATABASE_URL"], created_by="dev-user@example.test")
    with psycopg.connect(sample_app.config["DATABASE_URL"]) as connection:
        row = connection.execute(
            "SELECT id FROM export_files WHERE export_run_id = %s",
            (run_id,),
        ).fetchone()

    assert row is not None

    app = create_app(
        AppConfig(
            app_env="test",
            secret_key="test-secret",
            database_url=sample_app.config["DATABASE_URL"],
            allowed_emails=("allowed@example.test",),
            oidc_enabled=False,
            oidc_testing_profile={
                "sub": "dev-user",
                "email": "dev-user@example.test",
            },
            oidc_cookie_secure=False,
            llm_enabled=False,
        ),
        {"TESTING": True},
    )

    response = app.test_client().get(f"/export/files/{row[0]}")

    assert response.status_code == 403


def test_export_route_shows_previous_exports_and_failed_runs(sample_app: Flask) -> None:
    generate_budget_export(sample_app.config["DATABASE_URL"], created_by="dev-user@example.test")
    generate_budget_export(sample_app.config["DATABASE_URL"], created_by="dev-user@example.test")
    with psycopg.connect(sample_app.config["DATABASE_URL"]) as connection:
        with connection.transaction():
            connection.execute(
                """
                INSERT INTO export_runs (
                    export_type,
                    period_start,
                    period_end,
                    status,
                    started_at,
                    finished_at,
                    error_message
                )
                VALUES (
                    'budget_averages',
                    '2026-05-01',
                    '2026-05-31',
                    'failed',
                    now(),
                    now(),
                    'sample export failure'
                )
                """
            )

    response = sample_app.test_client().get("/export")
    body = response.get_data(as_text=True)

    assert response.status_code == 200
    assert body.count("budget-averages-2026-05.xlsx") == 2
    assert "failed: sample export failure" in body
    assert body.count("Download") == 2


def _extract_csrf(html: str) -> str:
    match = re.search(r'name="csrf_token" value="([^"]+)"', html)
    if match is None:
        raise AssertionError("CSRF token was not rendered")
    return match.group(1)
